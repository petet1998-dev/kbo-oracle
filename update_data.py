#!/usr/bin/env python3
"""
KBO Oracle — 통합 데이터 업데이트 도구 (관리자용)

엑셀 하나로 두 파일을 동시에 생성합니다:
  1. kbo_data.json        → 소개 웹사이트(GitHub Pages)용 (간단 형식)
  2. dataset_external.json → 진짜 엔진(Streamlit)용 (전체 형식)

사용법:
    python update_data.py KBO_전체데이터_입력용.xlsx

데이터 늘릴 때마다 이거 한 번 돌리고, 만들어진 두 파일을
GitHub에 덮어쓰기만 하면 사이트 + 앱 둘 다 업데이트됩니다.
"""

import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl 없음. 설치: pip install openpyxl")
    sys.exit(1)


# 엑셀 헤더 → 내부 키 (한글/영문 모두 인식)
COLUMN_MAP = {
    "선수명 (영문)": "name", "선수명": "name", "Name": "name",
    "데뷔년도": "debut_year",
    "직전리그": "prev_league", "직전 리그": "prev_league",
    "type_label": "type_label",
    "투타": "throws",
    "나이": "age",
    "G": "g", "GS": "gs", "IP": "ip",
    "K/9": "k9", "BB/9": "bb9", "HR/9": "hr9",
    "FIP": "fip", "평속 (mph)": "avg_velo", "평속": "avg_velo",
    "IVB (inch)": "ivb_inch", "IVB": "ivb_inch", "BABIP": "babip",
    "SwStr%": "swstr_pct", "CSW%": "csw_pct", "GB%": "gb_pct", "xFIP": "xfip",
    "직전2년 IP": "prev_2y_ip", "직전2년 ERA": "prev_2y_era", "직전2년 K/9": "prev_2y_k9",
    "구속 추세": "velo_trend", "IL 일수": "il_days_2y",
    "KBO WAR": "actual_war", "WAR": "actual_war",
    "구종": "pitch_types", "비고": "_note",
}

NUMERIC_KEYS = {"actual_war", "k9", "bb9", "hr9", "fip", "avg_velo", "ip",
                "swstr_pct", "csw_pct", "gb_pct", "xfip", "ivb_inch", "babip",
                "prev_2y_ip", "prev_2y_era", "prev_2y_k9", "velo_trend"}
INT_KEYS = {"debut_year", "age", "g", "gs", "il_days_2y"}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "N/A", "n/a", "null", "#N/A"):
        return None
    return s


def to_num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def classify(war):
    """WAR → 라벨 (3.5 이진, 도메인 기준)"""
    if war is None:
        return None
    return "Success" if war >= 3.5 else "Failure"


def convert(xlsx_path: str):
    path = Path(xlsx_path)
    if not path.exists():
        print(f"❌ 파일 없음: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌ 빈 시트")
        sys.exit(1)

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_idx = {COLUMN_MAP[h]: i for i, h in enumerate(header) if h in COLUMN_MAP}

    if "name" not in col_idx:
        print(f"❌ '선수명' 컬럼 못 찾음. 인식된 컬럼: {list(col_idx.keys())}")
        sys.exit(1)

    print(f"✅ 인식된 컬럼 {len(col_idx)}개: {list(col_idx.keys())}")

    skip_kw = {"FanGraphs 영문", "★필수", "필수", "KBO 첫 시즌", "자동 라벨", "Success/Marginal/Failure"}

    full_players = []   # 엔진용 (전체)
    web_players = []    # 웹용 (간단)

    for row in rows[1:]:
        name = clean(row[col_idx["name"]]) if col_idx["name"] < len(row) else None
        if not name or name in skip_kw or any(kw in str(name) for kw in skip_kw):
            continue

        rec = {"name": name}
        for key, idx in col_idx.items():
            if key in ("name", "_note") or idx >= len(row):
                continue
            val = clean(row[idx])
            if val is None:
                continue
            if key == "pitch_types":
                # "FB, SL, CH" → 리스트
                import re
                toks = [t.strip().upper() for t in re.split(r"[,\|/\s]+", val) if t.strip()]
                if toks:
                    rec[key] = toks
            elif key == "throws":
                rec[key] = val.upper()[:1]  # R/L
            elif key == "prev_league":
                rec[key] = val.upper()
            elif key in INT_KEYS:
                n = to_num(val)
                if n is not None:
                    rec[key] = int(n)
            elif key in NUMERIC_KEYS:
                n = to_num(val)
                if n is not None:
                    rec[key] = n
            else:
                rec[key] = val

        # 라벨 자동 (WAR 기준 3.5)
        war = rec.get("actual_war")
        if war is not None:
            rec["type_label"] = classify(war)
            # 엔진이 prev_ip도 쓰므로 ip 복사
            if "ip" in rec and "prev_ip" not in rec:
                rec["prev_ip"] = rec["ip"]

        rec["data_source"] = "user_excel"
        full_players.append(rec)

        # 웹용 간단 버전
        web_players.append({
            "name": name,
            "league": rec.get("prev_league", "?"),
            "war": war,
            "label": rec.get("type_label"),
            "k9": rec.get("k9"),
            "bb9": rec.get("bb9"),
        })

    # WAR 높은 순 정렬
    full_players.sort(key=lambda p: p.get("actual_war", -999), reverse=True)
    web_players.sort(key=lambda p: p.get("war", -999) if p.get("war") is not None else -999, reverse=True)

    # 저장
    base = path.parent
    web_path = base / "kbo_data.json"
    engine_path = base / "dataset_external.json"

    with open(web_path, "w", encoding="utf-8") as f:
        json.dump(web_players, f, ensure_ascii=False, indent=2)
    with open(engine_path, "w", encoding="utf-8") as f:
        json.dump(full_players, f, ensure_ascii=False, indent=2)

    # 통계
    total = len(full_players)
    labeled = [p for p in full_players if p.get("actual_war") is not None]
    success = sum(1 for p in labeled if p["actual_war"] >= 3.5)
    with_raw = sum(1 for p in full_players if p.get("k9") is not None)

    print()
    print("=" * 55)
    print("✅ 두 파일 생성 완료")
    print("=" * 55)
    print(f"  📊 총 선수: {total}명")
    print(f"  📊 WAR 있음: {len(labeled)}명 (Success {success} / Failure {len(labeled)-success})")
    print(f"  📊 raw 스탯(K/9) 있음: {with_raw}명")
    print()
    print(f"  1️⃣ {web_path.name}        → GitHub Pages(소개 사이트)에 덮어쓰기")
    print(f"  2️⃣ {engine_path.name} → Streamlit(진짜 앱) GitHub에 덮어쓰기")
    print()
    print("  ⚠️ 주의: raw 스탯 없는 선수가 많으면 모델 정확도가 오히려")
    print("     떨어질 수 있음. K/9·BB/9·직전리그는 꼭 채우세요.")

    # raw 결측 경고
    no_raw = total - with_raw
    if no_raw > total * 0.3:
        print()
        print(f"  🔴 경고: {no_raw}명이 raw 스탯 결측입니다. 모델 학습 품질 저하 위험.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python update_data.py <엑셀파일.xlsx>")
        print("예시:   python update_data.py KBO_전체데이터_입력용.xlsx")
        sys.exit(1)
    convert(sys.argv[1])
