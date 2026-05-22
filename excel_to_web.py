#!/usr/bin/env python3
"""
KBO Oracle — 엑셀 → 웹사이트 JSON 변환기

사용법:
    python excel_to_web.py KBO_전체데이터_입력용.xlsx

결과:
    kbo_data.json 생성 → index.html과 같은 폴더에 넣으면 사이트가 자동으로 읽음

승수님이 엑셀 채울 때마다 이 스크립트 한 번만 돌리면 사이트 전체가 업데이트됩니다.
"""

import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl 없음. 설치: pip install openpyxl")
    sys.exit(1)


# 엑셀 헤더 → JSON 키 매핑 (한글/영문 모두 인식)
COLUMN_MAP = {
    "선수명 (영문)": "name", "선수명": "name", "Name": "name",
    "데뷔년도": "debut_year",
    "직전리그": "league", "직전 리그": "league",
    "type_label": "label",
    "K/9": "k9", "BB/9": "bb9", "HR/9": "hr9",
    "FIP": "fip", "평속 (mph)": "velo", "평속": "velo",
    "IP": "ip",
    "SwStr%": "swstr", "CSW%": "csw", "GB%": "gb", "xFIP": "xfip",
    "KBO WAR": "war", "WAR": "war",
}

# 숫자로 변환할 키
NUMERIC_KEYS = {"war", "k9", "bb9", "hr9", "fip", "velo", "ip",
                "swstr", "csw", "gb", "xfip", "debut_year"}


def clean_value(v):
    """빈 값 정리"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "—", "N/A", "n/a", "null", "#N/A"):
        return None
    return s


def to_number(v):
    """숫자 변환 (실패하면 None)"""
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def classify(war):
    """WAR → 라벨 (3.5 이진 기준, 승수님 도메인)"""
    if war is None:
        return None
    return "Success" if war >= 3.5 else "Failure"


def convert(xlsx_path: str):
    path = Path(xlsx_path)
    if not path.exists():
        print(f"❌ 파일 없음: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)  # 수식 결과값 읽기
    ws = wb.active  # 첫 시트 (입력 시트)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌ 빈 시트")
        sys.exit(1)

    # 헤더 찾기 (첫 행)
    header = [str(c).strip() if c else "" for c in rows[0]]
    print(f"📋 헤더: {header[:6]}...")

    # 헤더 인덱스 매핑
    col_idx = {}
    for i, h in enumerate(header):
        if h in COLUMN_MAP:
            col_idx[COLUMN_MAP[h]] = i

    if "name" not in col_idx:
        print("❌ '선수명' 컬럼을 못 찾음. 헤더 확인 필요")
        print(f"   인식된 컬럼: {list(col_idx.keys())}")
        sys.exit(1)

    print(f"✅ 인식된 컬럼: {list(col_idx.keys())}")

    # 데이터 행 파싱 (도움말/안내 행 스킵)
    players = []
    skip_keywords = {"FanGraphs 영문", "★필수", "필수", "KBO 첫 시즌", "자동 라벨"}

    for row in rows[1:]:
        name_raw = row[col_idx["name"]] if col_idx["name"] < len(row) else None
        name = clean_value(name_raw)
        if not name:
            continue
        # 도움말 행 스킵
        if name in skip_keywords or any(kw in str(name) for kw in skip_keywords):
            continue

        player = {"name": name}
        for key, idx in col_idx.items():
            if key == "name":
                continue
            if idx >= len(row):
                continue
            val = clean_value(row[idx])
            if val is None:
                continue
            if key in NUMERIC_KEYS:
                num = to_number(val)
                if num is not None:
                    player[key] = num
            else:
                player[key] = val

        # 라벨 자동 계산 (WAR 기준)
        if "war" in player:
            player["label"] = classify(player["war"])

        players.append(player)

    # WAR 높은 순 정렬
    players.sort(key=lambda p: p.get("war", -999), reverse=True)

    # JSON 저장 (사이트와 같은 폴더)
    out_path = path.parent / "kbo_data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(players, f, ensure_ascii=False, indent=2)

    # 통계
    total = len(players)
    labeled = [p for p in players if p.get("war") is not None]
    success = sum(1 for p in labeled if p["war"] >= 3.5)
    failure = len(labeled) - success
    with_k9 = sum(1 for p in players if p.get("k9") is not None)

    print()
    print("=" * 50)
    print(f"✅ 변환 완료: {out_path}")
    print("=" * 50)
    print(f"  총 선수: {total}명")
    print(f"  WAR 있음: {len(labeled)}명")
    print(f"    - Success (≥3.5): {success}명")
    print(f"    - Failure (<3.5): {failure}명")
    print(f"  K/9 데이터 있음: {with_k9}명")
    print()
    print("📌 다음 단계:")
    print(f"  1. {out_path.name} 파일을 index.html과 같은 폴더에 두세요")
    print("  2. 사이트 새로고침하면 자동으로 새 데이터가 반영됩니다")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python excel_to_web.py <엑셀파일.xlsx>")
        print("예시:   python excel_to_web.py KBO_전체데이터_입력용.xlsx")
        sys.exit(1)
    convert(sys.argv[1])
